# Importē visas nepieciešamās bibliotēkas. Ja bibliotēkas nestrādā, tad ir jāizpild kods "pip install selenium".
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook, load_workbook

# Lietotājam dod ievadīt meklētās filmas nosaukumu.
print("Ievadiet filmas nosaukumu: ")
film = input()

# Nodefinē mainīgo driver, lai programma spētu kontrolēt pārlūkprogrammu Google Chrome.
service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service = service, options = option)



# Tiek definēta web adrese, kur tiek meklēta informācija par filmas reģionu.
url = "https://www.google.com/"

# Atver mājaslapu, uz kuru aizved adrese.
driver.get(url)
time.sleep(0.5)

# Noraida Google sīkfailus.
decline_cookies = driver.find_element(By.ID, "W0wltc")
decline_cookies.click()
time.sleep(0.1)

# Ievada meklējamo frāzi logā un to sameklē ar Google.
search_box = driver.find_element(By.ID, "APjFqb")
search = film + " netflix"
search_box.send_keys(search, Keys.ENTER)
time.sleep(1)

# Atrod filmas netflix saites adresi un aiztaisa Google Chrome.
find = driver.find_element(By.XPATH, "//a[@jsname='UWckNb']")
link = find.get_attribute("href")
driver.close()

# Definē mainīgos, kādi tiks izmantoti, lai atrastu valsti, kurā ir filma ir pieejama.
b = []
b = link 
if (len(b) > 29):
page_name = b[12] + b[13] + b[14] + b[15] + b[16] + b[17] + b[18] 
region = b[24] + b[25]

# Izvelk Netflix reģiona kodu no saites adreses un atrod atbilstošo valsti Excel tabulā. Izvada rezultātu.
  if (page_name == "netflix" and (region + b[26] + b[27] + b[28] + b[29]) != "search"):
    if ((region + b[26] + b[27] + b[28]) == "title"):
      print("Šī filma ir pieejama Jūsu pašreizējā reģionā.")
    else:
      wb = load_workbook('CountryCodes.xlsx')
      ws = wb.active
      for row in range (1, 250):
        country_code=(ws['b' + str(row)].value)
        country_name=(ws['a' + str(row)].value)
        if (country_code == region):
          print("Filma ir pieejama šādā reģionā: " + country_name)
          row = row - 1
        if (row == 250):
          print("Šī filma nav pieejama vietnē Netflix.")
  else:
    print("Šī filma nav pieejama vietnē Netflix.")
else:
  print("Šī filma nav pieejama vietnē Netflix.")
